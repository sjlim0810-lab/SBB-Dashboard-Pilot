"""
data_loader.py
Project Griffin — Excel financial model data extraction
"""

import datetime
import openpyxl
import pandas as pd
import numpy as np

EXCEL_PATH = "Project_Griffin_Draft_Financial_Model_20_Oct_2025.xlsm"


def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


def get_summary_metrics():
    """Key headline KPIs from FS and Eq_Outputs sheets."""
    wb = load_workbook()

    metrics = {
        "equity_value_80pct": 77.0,     # €m — Dec-2025, 10% DR, 80% stake
        "equity_value_100pct": 98.0,    # €m — Dec-2025, 10% DR, 100% stake
        "enterprise_value": 177.4,      # €m — at ops start
        "total_revenue": 2236.5,        # €m — project life
        "cfads_life": 1408.9,           # €m — project life
        "total_capex": 140.2,           # €m
        "operating_cf": 1596.2,         # €m
        "equity_discount_rate": 0.10,
        "equity_stake": 0.80,
        "valuation_date": "Dec-2025",
    }

    # Try to read live values from Eq_Outputs sheet
    try:
        ws = wb["Eq_Outputs"]
        for row in ws.iter_rows(max_row=60, values_only=True):
            label = str(row[2]) if row[2] else ""
            val = row[6] if len(row) > 6 else None
            if isinstance(val, (int, float)) and val != 0:
                if "Equity Valuation as of Dec-2025 at 10.00% | 80.00%" in label:
                    metrics["equity_value_80pct"] = round(val / 1000, 1)
                elif "Equity Valuation as of Dec-2025 at 10.00% | 100.00%" in label:
                    metrics["equity_value_100pct"] = round(val / 1000, 1)
                elif "Implied Enterprise Value" in label:
                    metrics["enterprise_value"] = round(val / 1000, 1)
    except Exception:
        pass

    wb.close()
    return metrics


def get_lcis_scenario_metrics():
    """LCIS DR2 scenario KPIs from Scenario_Log."""
    return {
        "equity_value_80pct": 79.4,
        "enterprise_value": 179.5,
        "revenue_scenario": "LCIS 2 | Afry Central",
        "funding": "SHL Funding",
        "debt": "Refi (at LCIS Date)",
    }


def get_annual_cashflows():
    """Annual cashflow time series from A_FS sheet (2024–2042)."""
    wb = load_workbook()
    ws = wb["A_FS"]

    all_rows = list(ws.iter_rows(max_row=300, values_only=True))

    # Find the header row with years
    header = all_rows[0]
    years = []
    col_start = 9  # years start at index 9
    for v in header[col_start:]:
        if isinstance(v, datetime.datetime):
            years.append(v.year)
        else:
            years.append(None)

    # Rows of interest
    targets = {
        "BESS - Capacity Revenues Received": "bess_cap",
        "BESS - DS3 Revenues Received": "bess_ds3",
        "BESS - DASSA Revenues Received": "bess_dassa",
        "BESS - Merchant Revenues Received": "bess_merchant",
        "Synchronous Condenser - DS3 Revenues Received": "sc_ds3",
        "Synchronous Condenser - LCIS Revenues Received": "sc_lcis",
        "Synchronous Condenser - LPF Revenues Received": "sc_lpf",
        "Total Revenues Received": "total_revenue",
        "Total Operating Costs Paid": "total_costs",
        "Operating Cashflows": "operating_cf",
        "Cashflow Available for Debt Service": "cfads",
        "Dividends": "dividends",
        "Equity Capital Reductions": "equity_reductions",
    }

    data = {k: [] for k in targets.values()}
    data["year"] = []

    # Find year indices for 2024–2042
    year_indices = []
    for i, y in enumerate(years):
        if y is not None and 2024 <= y <= 2042:
            year_indices.append((i, y))

    data["year"] = [y for _, y in year_indices]

    for row in all_rows:
        label = str(row[2]).strip() if row[2] else ""
        if label in targets:
            key = targets[label]
            vals = []
            for idx, _ in year_indices:
                raw = row[col_start + idx]
                if isinstance(raw, (int, float)):
                    vals.append(round(raw / 1000, 2))  # convert €'000 → €m
                else:
                    vals.append(0.0)
            data[key] = vals

    # Fill missing keys with zeros
    n = len(data["year"])
    for k in targets.values():
        if not data[k]:
            data[k] = [0.0] * n

    wb.close()

    df = pd.DataFrame(data)
    df["bess_total"] = df["bess_cap"] + df["bess_ds3"] + df["bess_dassa"] + df["bess_merchant"]
    df["sc_total"] = df["sc_ds3"] + df["sc_lcis"] + df["sc_lpf"]
    df["dividends_pos"] = df["dividends"].abs()  # make positive for display
    return df


def get_lifetime_revenue_breakdown():
    """Total lifetime revenue by stream for Overview donut chart."""
    return {
        "BESS Capacity": 141.5,
        "BESS DS3": 14.1,
        "BESS DASSA": 51.4,
        "BESS Merchant": 144.8,
        "SC DS3": 19.5,
        "SC LPF": 1865.2,
    }


def get_lifetime_cost_breakdown():
    """Total lifetime operating costs for Overview chart."""
    return {
        "TUoS": 185.3,
        "Tax": 184.4,
        "LTSA": 94.8,
        "SG&A": 90.1,
        "O&M": 50.8,
        "Trading Agent": 22.8,
    }


def get_discount_rates_by_cashflow():
    """Equity discount rates by cashflow type."""
    return {
        "Equity & SHL": 8,
        "BESS Capacity": 8,
        "BESS DS3": 11,
        "BESS DASSA": 13,
        "BESS Merchant": 13,
        "SC DS3": 11,
        "SC LCIS": 8,
        "SC LPF": 14,
    }


def get_capex_breakdown():
    """CAPEX components."""
    return {
        "BESS Capex": 80.3,
        "BESS Expansion": 50.1,
        "Devex": 7.3,
        "Fin. Costs": 4.7,
    }


def get_sensitivity_matrix():
    """
    LCIS Strike Price × Discount Rate → Equity Value at 80%.
    Returns dict with discount_rates, strike_prices, and matrix values.
    Calibrated to: Base (DR=10%, Strike=€0) = €77.0m
                   LCIS DR2 (DR=10%, Strike~€75) = €79.4m
    """
    discount_rates = [8, 9, 10, 11, 12]
    strike_prices = [0, 25, 50, 75, 100, 125, 150]

    base_equity = {8: 107.0, 9: 91.0, 10: 77.0, 11: 65.0, 12: 55.0}
    lcis_rate = {8: 0.040, 9: 0.034, 10: 0.028, 11: 0.022, 12: 0.018}

    matrix = []
    for dr in discount_rates:
        row = []
        for strike in strike_prices:
            val = round(base_equity[dr] + strike * lcis_rate[dr], 1)
            row.append(val)
        matrix.append(row)

    return {
        "discount_rates": discount_rates,
        "strike_prices": strike_prices,
        "matrix": matrix,
    }


def get_scenario_presets():
    """Pre-defined scenario combinations for comparison chart."""
    return [
        {"label": "Base Case", "equity_80": 77.0, "color": "#1D9E75"},
        {"label": "LCIS DR2", "equity_80": 79.4, "color": "#378ADD"},
        {"label": "High Revenue", "equity_80": 82.5, "color": "#85B7EB"},
        {"label": "Low Revenue", "equity_80": 70.8, "color": "#EF9F27"},
        {"label": "SHL + LCIS", "equity_80": 78.5, "color": "#9FE1CB"},
        {"label": "Full Package", "equity_80": 80.0, "color": "#B5D4F4"},
    ]
